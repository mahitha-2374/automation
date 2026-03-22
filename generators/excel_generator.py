"""
Excel T3 Generator
Safely fills T3 XLSM template preserving formulas
"""

from openpyxl import load_workbook
import datetime
import os


class ExcelGenerator:
    """Generate T3 Excel files while preserving formulas"""

    def __init__(self):
        """Initialize Excel generator"""
        self.sheet_mappings = {
            "user_role": "User_role_resource",
            "user_account": "User_Account_lookup",
            "role_resource": "Role_Resource",
            "role_resource_lookup": "Role_resource_lookup",
            "gsi_user_role": "gsi_user-role-resource",
            "control": "gsi_user-role-resource-cntrl"
        }

    def generate(self, template_path, output_path, data):
        """
        Fill T3 template with processed data
        
        Args:
            template_path: Path to T3_template.xlsm
            output_path: Where to save output
            data: Processed IAM data with gsi_id
        """
        if not os.path.exists(template_path):
            # Create a simple template structure if it doesn't exist
            self._create_basic_template(template_path)

        wb = load_workbook(template_path, keep_vba=True)

        # Fill Summary/GSI Sheet
        self._fill_summary_sheet(wb, data)

        # Fill User Role Resource sheet
        self._fill_user_role_resource(wb, data)

        # Fill Role Resource sheet
        self._fill_role_resource(wb, data)

        # Fill Role Resource Lookup sheet
        self._fill_role_resource_lookup(wb, data)

        # Fill User Account Lookup sheet
        self._fill_user_account_lookup(wb, data)

        # Fill Control Sheet
        self._fill_control_sheet(wb, data)

        wb.save(output_path)

    def _fill_summary_sheet(self, wb, data):
        """Fill Summary sheet with GSI and metadata"""
        try:
            ws = wb["Summary"]
        except:
            ws = wb.create_sheet("Summary", 0)  # Insert as first sheet

        gsi_id = data.get("gsi_id", "NOT_PROVIDED")
        
        # Add summary information
        ws["A1"] = "Application Summary"
        
        ws["A3"] = "GSI ID:"
        ws["B3"] = gsi_id
        
        ws["A4"] = "Total Users:"
        ws["B4"] = len(data.get("users", []))
        
        ws["A5"] = "Total Roles:"
        ws["B5"] = len(data.get("roles", []))
        
        ws["A6"] = "Total Entitlements:"
        ws["B6"] = len(data.get("entitlements", []))
        
        ws["A7"] = "Generated Date:"
        ws["B7"] = str(datetime.date.today())
        
        ws["A8"] = "Generated Time:"
        ws["B8"] = str(datetime.datetime.now().strftime("%H:%M:%S"))
        
        # User-Role Mappings
        user_role_count = sum(len(user["roles"]) for user in data.get("users", []))
        ws["A9"] = "User-Role Mappings:"
        ws["B9"] = user_role_count

    def _fill_user_role_resource(self, wb, data):
        """Fill User_role_resource sheet"""
        try:
            ws = wb["User_role_resource"]
        except:
            ws = wb.active
            ws.title = "User_role_resource"

        row = 2

        for user in data.get("users", []):
            for role in user.get("roles", []):
                ws.cell(row=row, column=1).value = user["user_id"]
                ws.cell(row=row, column=6).value = role
                row += 1

    def _fill_role_resource(self, wb, data):
        """Fill Role_Resource sheet"""
        try:
            ws = wb["Role_Resource"]
        except:
            ws = wb.create_sheet("Role_Resource")

        row = 2

        for role in data.get("roles", []):
            for ent in role.get("entitlements", []):
                ws.cell(row=row, column=1).value = role["role_name"]
                ws.cell(row=row, column=2).value = ent
                row += 1

    def _fill_role_resource_lookup(self, wb, data):
        """Fill Role Resource Lookup sheet"""
        try:
            ws = wb["Role_resource_lookup"]
        except:
            ws = wb.create_sheet("Role_resource_lookup")

        row = 2

        # Add roles first
        for role in data.get("roles", []):
            ws.cell(row=row, column=1).value = role["role_name"]
            row += 1

        # Add entitlements
        for ent in data.get("entitlements", []):
            ws.cell(row=row, column=5).value = ent["resource_name"]
            ws.cell(row=row, column=6).value = ent.get("description", "")
            row += 1

    def _fill_user_account_lookup(self, wb, data):
        """Fill User Account Lookup sheet"""
        try:
            ws = wb["User_Account_lookup"]
        except:
            ws = wb.create_sheet("User_Account_lookup")

        row = 2

        for user in data.get("users", []):
            ws.cell(row=row, column=1).value = user["user_id"]
            ws.cell(row=row, column=2).value = user.get("account_status", "ACTIVE")
            row += 1

    def _fill_control_sheet(self, wb, data):
        """Fill control sheet with row count and run date"""
        try:
            ws = wb["gsi_user-role-resource-cntrl"]
        except:
            ws = wb.create_sheet("gsi_user-role-resource-cntrl")

        user_role_count = sum(
            len(user["roles"]) for user in data.get("users", [])
        )

        ws.cell(row=2, column=1).value = user_role_count
        ws.cell(row=2, column=2).value = str(datetime.date.today())

    def _create_basic_template(self, path):
        """Create a basic template if it doesn't exist"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "User_role_resource"

        # Add headers
        ws["A1"] = "UserID"
        ws["B1"] = "Role_Name"

        wb.save(path)
