"""
Create a sample DSAP template for demonstration
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "DSAP"

# Set column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 20

# Header style
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
header_alignment = Alignment(horizontal="center", vertical="center")

# Add title
ws['A1'] = "DSAP - Digital System Access Plan"
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:F1')

# Add metadata
ws['A2'] = f"GSI ID: {{{{gsi_id}}}}"
ws['A3'] = f"Export Date: {{{{export_date}}}}"
ws['A4'] = f"Export Time: {{{{export_time}}}}"

# Add summary section
ws['A6'] = "Summary"
ws['A6'].font = Font(bold=True, size=12)

ws['A7'] = "Total Users:"
ws['B7'] = "{{total_users}}"
ws['A8'] = "Total Roles:"
ws['B8'] = "{{total_roles}}"

# Add column headers for users/roles
row = 10
headers = ["User ID", "Role Name", "Status", "Manager", "Module", "Description"]

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Add sample rows (these will be filled with data)
row = 11
for i in range(5):
    ws.cell(row=row+i, column=1).value = "{{user_id}}" if i == 0 else f"[USER_{i}]"
    ws.cell(row=row+i, column=2).value = "{{role_name}}" if i == 0 else f"[ROLE_{i}]"
    ws.cell(row=row+i, column=3).value = "{{status}}" if i == 0 else "ACTIVE"
    ws.cell(row=row+i, column=4).value = "{{manager}}" if i == 0 else "[MANAGER]"
    ws.cell(row=row+i, column=5).value = "{{module}}" if i == 0 else "[MODULE]"
    ws.cell(row=row+i, column=6).value = "{{description}}" if i == 0 else "[DESC]"

# Add borders to data cells
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in range(10, 16):
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = thin_border

# Save template
template_path = Path("templates") / "DSAP_template.xlsx"
template_path.parent.mkdir(exist_ok=True)
wb.save(template_path)

print(f"✓ Sample DSAP template created: {template_path}")
print("\nTemplate placeholders detected:")
print("  • {{gsi_id}} - System GSI ID")
print("  • {{export_date}} - Export date")
print("  • {{export_time}} - Export time")
print("  • {{total_users}} - Total user count")
print("  • {{total_roles}} - Total role count")
print("  • {{user_id}}, {{role_name}}, {{status}}, {{manager}}, {{module}}, {{description}}")
print("\nHint: Just provide your own templates and they'll work automatically!")
